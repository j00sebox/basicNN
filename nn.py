import math
import numpy as np
import xlwt 
from openpyxl import Workbook
import pandas as pd
import os

# sigmod function for neuron activation values
def sigmoid(x):
    # stable version of sigmoid to prevent overflows
    result = np.where(x >= 0, 
                    1 / (1 + np.exp(-x)), 
                    np.exp(x) / (1 + np.exp(x)))
    return result

# softmax fucntion used for activation when using cross-entopy loss
def softmax(x):
    # stable version of softmax, more negative numbers will be pushed to zero
    exps = np.exp(x - np.max(x))
    return exps / np.sum(exps)

# Class to instantiate network structure
class NeuralNetwork(object):

    # takes in topology of neural network
    # second parameter is an array of hidden layer nodes
    def __init__(self, inputNodes, hiddenLayer, outputNodes, outputLabels, learningRate = 0.1, weightDecay = 0.01):
        # initialize network structure
        self.i = inputNodes
        self.o = outputNodes
        self.hiddenLayer = hiddenLayer
        self.weight_matrix_list = []
        self.bias_matrix_list = []
        self.activation_matrix_list = []

        # set learning rate
        self.lr = learningRate

        # set weight decay
        self.wd = weightDecay

        # labels for what each output neuron represents
        self.labels = outputLabels

        # add input weights to the maxtrix list
        weights_1 = np.random.uniform(-1.01, 1.01, size=(self.hiddenLayer[0], inputNodes))
        self.weight_matrix_list.append(weights_1)

        # add input biases to the maxtrix list
        bias_1 = np.random.uniform(-1.01, 1.01, size=(self.hiddenLayer[0], 1))
        self.bias_matrix_list.append(bias_1)
        
        # initialize weights and biases for hidden layers
        if len(hiddenLayer) > 1:
            for i in range(1, len(self.hiddenLayer)):
                weights = np.random.uniform(-1.01, 1.01, size=(self.hiddenLayer[i], self.hiddenLayer[i - 1]))
                self.weight_matrix_list.append(weights)

                bias = np.random.uniform(-1.01, 1.01, size=(self.hiddenLayer[i], 1))
                self.bias_matrix_list.append(bias)

        # add output weights to maxtrix list
        weights_2 = np.random.uniform(-1.01, 1.01, size=(outputNodes, self.hiddenLayer[-1]))
        self.weight_matrix_list.append(weights_2)

        # add output biases to matrix list
        bias_2 = np.random.uniform(-1.01, 1.01, size=(outputNodes, 1))
        self.bias_matrix_list.append(bias_2)

    # returns the neural network's guess for the correct output based on the data
    def guess(self, data):
        
        # get output activations
        output = self.compute_neurons(data).tolist()

        # take the max activation as the guess
        guess = output.index(max(output))

        # return guess
        return self.labels[guess]
        
    # feed activations through network 
    # returns the activations of the output layer
    def compute_neurons(self, data):

        self.InputData = np.asarray(data)

        # append inputs as first element in activation list
        self.activation_matrix_list.append(self.InputData)

        # calculate activation for hidden neurons
        self.activation_matrix_list.append(sigmoid(self.weight_matrix_list[0].dot(self.InputData) + self.bias_matrix_list[0]))

        # cycle through hidden layers
        if len(self.hiddenLayer) > 1:
            for i in range(1, len(self.hiddenLayer)):
                self.activation_matrix_list.append(sigmoid(self.weight_matrix_list[i].dot(self.activation_matrix_list[i]) + self.bias_matrix_list[i]))

        # calculate activation for output neurons
        self.activation_matrix_list.append(softmax(self.weight_matrix_list[-1].dot(self.activation_matrix_list[-1]) + self.bias_matrix_list[-1]))
        
        # copy list over to be used for training
        self.a = self.activation_matrix_list

        # set list as empty before next feedforward
        self.activation_matrix_list = []

        # return best guess
        return self.a[-1]

    # uses backpropagation to optimize the weights and biases
    def train(self, inputs, targets):
        # start by letting the network guess
        result = self.compute_neurons(inputs)

        #self.mse_backprop(result, targets)

        # get error for one case
        cross_ent_err = targets - result
        self.cross_entropy_backprop(result, targets, cross_ent_err)

    def train_batch(self, inputs, targets, tItr):
        cross_ent_err = 0
        out = 0
        # get average error over training batch
        for i in range(0, tItr):
            out = self.compute_neurons(inputs[i])
            cross_ent_err += targets[i] - out

        self.cross_entropy_backprop(out, targets, cross_ent_err)

    def cross_entropy(self, x, y):
        loss = -x*np.log(y)
        return loss
            
    def cross_entropy_backprop(self, result, targets, err):
        # set as empty before calculations
        backprop_error = []
        weight_list = []
        bias_list = []
        a_list = []

        # convert targets to matrix
        targets = np.asarray(targets)

        backprop_error.append(err)

        # create new lists with reversed order
        weight_list = self.weight_matrix_list[::-1]
        bias_list = self.bias_matrix_list[::-1]
        a_list = self.a[::-1]

        # backpropagate error to the hidden layer
        for i in range(0, len(self.weight_matrix_list)):
            backprop_error.append(weight_list[i].transpose().dot(backprop_error[i]))

        # calculate gradient for cross entropy
        gradient = backprop_error[0]*self.lr
        delta_weights = gradient.dot(a_list[1].transpose())
        weight_list[0] += delta_weights

        bias_list[0] += gradient

        # loop through and calculate the gradient decsent
        for i in range(1, len(self.a) - 1):
            # calculate and add adjustments for the weights
            gradient = self.lr*2*a_list[i]*(1 - a_list[i])*backprop_error[i] 
            delta_weights = gradient.dot(a_list[i+1].transpose()) - 2*self.wd*weight_list[i]

            # apply delta to weights
            weight_list[i] += delta_weights 

            # adjust bias accordingly
            bias_list[i] += gradient
        
        # set the weights and biases as the adjusted values
        self.weight_matrix_list = weight_list[::-1]
        self.bias_matrix_list = bias_list[::-1]

    def mse(self, x, y):
        error = np.power((x - y), 2)
        return error

    def mse_backprop(self, result, targets):
        # set as empty before calculations
        backprop_error = []
        weight_list = []
        bias_list = []
        a_list = []

        # compute error of guess
        error = targets - result
        backprop_error.append(error)

        # create new lists with reversed order
        weight_list = self.weight_matrix_list[::-1]
        bias_list = self.bias_matrix_list[::-1]
        a_list = self.a[::-1]

        # backpropagate error to the hidden layer
        for i in range(0, len(self.weight_matrix_list)):
            backprop_error.append(weight_list[i].transpose().dot(backprop_error[i]))

        # loop through and calculate the gradient decsent
        for i in range(0, len(self.a) - 1):
            # calculate and add adjustments for the weights
            gradient = self.lr*2*a_list[i]*(1 - a_list[i])*backprop_error[i]
            delta_weights = gradient.dot(a_list[i+1].transpose())
            weight_list[i] += delta_weights

            # adjust bias accordingly
            bias_list[i] += gradient
        
        # set the weights and biases as the adjusted values
        self.weight_matrix_list = weight_list[::-1]
        self.bias_matrix_list = bias_list[::-1]

    def store_weights_and_biases(self):
        # if file exists remove
        if os.path.exists('weights.xlsx'):
            os.remove('weights.xlsx')

        if os.path.exists('biases.xlsx'):
            os.remove('biases.xlsx')

        # create workbook for weights
        w = Workbook() 

        # create workbook for biases
        b = Workbook()

        # loop through matrix list elements
        for i in range(0, len(self.weight_matrix_list)):
            # add_sheet is used to create sheet. 
            sheet = w.create_sheet('Sheet' + str(i))
            for j in range(0, len(self.weight_matrix_list[i])):     
                for x in range(0, len(self.weight_matrix_list[i][j])):
                    sheet.cell(j + 2, x + 1).value = self.weight_matrix_list[i][j][x]
        
        # save weights in file
        w.save('weights.xlsx')

        # loop through matrix list elements
        for i in range(0, len(self.bias_matrix_list)):
            # add_sheet is used to create sheet. 
            sheet = b.create_sheet('Sheet' + str(i))
            for j in range(0, len(self.bias_matrix_list[i])):
                for x in range(0, len(self.bias_matrix_list[i][j])):
                    sheet.cell(j + 2, 2).value = self.bias_matrix_list[i][j][x]

        # save biases in file
        b.save('biases.xlsx')

    def load_weights_and_biases(self):
        for i in range(0, len(self.weight_matrix_list)):
            df = pd.read_excel('weights.xlsx', sheet_name='Sheet' + str(i))
            for j in range(0, len(self.weight_matrix_list[i])):
                for x in range(0, len(self.weight_matrix_list[i][j])):
                    self.weight_matrix_list[i][j][x] = df.iat[j, x]
        
        for i in range(0, len(self.bias_matrix_list)):
            df = pd.read_excel('biases.xlsx', sheet_name='Sheet' + str(i))
            for j in range(0, len(self.bias_matrix_list[i])):
                for x in range(0, len(self.bias_matrix_list[i][j])):
                    self.bias_matrix_list[i][j][x] = df.iat[j, 1]
        